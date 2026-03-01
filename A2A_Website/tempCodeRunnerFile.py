from flask import Flask, request, render_template, send_file, redirect, url_for, jsonify
import os
import pandas as pd  
import uuid 
from flask_cors import CORS
# ★ 이 줄이 정확해야 합니다
from validation.validation_runner import run_validation
import json
import sys
from win32com.client import makepy, Dispatch
from validation.pipelines import (
    run_country_validation,
    get_all_step_functions,
    load_country_pipelines,
    save_country_pipeline,
    PIPELINE_PATH
)
# ----[추가]----
import openpyxl
# --------------

# ***** [NEW IMPORTS FOR PIVOT] *****
import pythoncom
import win32com.client as win32
import shutil
from win32com.client import gencache
# ***********************************
import time
import pywintypes

app = Flask(__name__)
CORS(app)  # 모든 출처에 대해 CORS 허용
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # app.py의 실제 위치

# 폴더 상수
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = os.path.join(app.root_path, "output")   # ← 루트 절대경로
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def safe_dispatch(prog_id, retries=3, delay=1):
    """
    Excel COM 객체를 안전하게 생성하는 함수.
    - Excel이 '바쁘다'면서 RPC_E_CALL_REJECTED 에러를 던질 때 재시도
    - retries: 최대 재시도 횟수
    - delay: 재시도 사이 대기 시간 (초)
    """
    for i in range(retries):
        try:
            return win32.gencache.EnsureDispatch(prog_id)
        except pywintypes.com_error as e:
            if e.hresult == -2147418111:  # RPC_E_CALL_REJECTED
                print(f"[WARN] Excel busy, retry {i+1}/{retries}...")
                time.sleep(delay)
                continue
            raise
    raise RuntimeError("Excel COM dispatch failed after retries")

# ***** [NEW: PIVOT HELPER FUNCTIONS] *****
def reset_com_cache_if_broken():
    try:
        _ = gencache.GetModuleForProgID('Excel.Application')
    except Exception:
        gen_py_path = gencache.GetGeneratePath()
        shutil.rmtree(gen_py_path, ignore_errors=True)
        gencache._cached_modules = {}

        # Also clear imported COM wrappers
        bad_modules = [m for m in sys.modules if m.startswith("win32com.gen_py.")]
        for m in bad_modules:
            del sys.modules[m]

        # Force regenerate Excel typelib
        gencache.EnsureModule(
            '{00020813-0000-0000-C000-000000000046}', 0, 1, 9
        )
    return win32.gencache.EnsureDispatch('Excel.Application')

def add_pivot_to_workbook(xlsx_path, data_sheet="Data_All"):
    """
    기존 통합파일(xlsx_path)에 PivotTable 시트를 추가한다.
    데이터는 `data_sheet` 시트 전체를 범위로 사용.
    """
    pythoncom.CoInitialize()
    makepy.GenerateFromTypeLibSpec("Microsoft Excel 16.0 Object Library")

    # 안전한 Excel 객체 생성
    excel = safe_dispatch("Excel.Application")
    excel.DisplayAlerts = False
    excel.Visible = False

    print("Excel object:", excel)
    #excel.Visible = False

    wb = excel.Workbooks.Open(xlsx_path)
    print(wb)
    ws_data  = wb.Sheets(data_sheet)
    ws_pivot = wb.Sheets.Add(After=ws_data)
    ws_pivot.Name = 'PivotTable'

# UsedRange 활용해 마지막 데이터 셀 감지
    used_range = ws_data.UsedRange
    first_row = used_range.Row
    first_col = used_range.Column
    last_row = first_row + used_range.Rows.Count - 1
    last_col = first_col + used_range.Columns.Count - 1
    data_range = ws_data.Range(ws_data.Cells(first_row, first_col), ws_data.Cells(last_row, last_col))

    # ----- Pivot 생성 -----
    pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)  # 1=xlDatabase
    pivot_table = pivot_cache.CreatePivotTable(
        TableDestination=ws_pivot.Cells(1, 1),
        TableName='MyPivotTable'
    )

    # ----- 레이아웃 -----
    pivot_table.PivotFields('Country').Orientation = 1          # xlRowField
    pivot_table.PivotFields('Carrier').Orientation = 1
    pivot_table.PivotFields('Carrier').Position = 2

    # ★ 추가: Remarks 를 Page-Filter 로
    pivot_table.PivotFields('Remarks').Orientation = 3 # xlPageField

        # ***** [NEW: Target Month] 컬럼을 Column 필드로 배치 *****
    try:
        pivot_table.PivotFields('Target Month').Orientation = 2  # xlColumnField
    except Exception as _e:
        # 필드가 없거나 읽기 실패 시 피벗 생성만 계속 진행
        pass
    # *******************************************************

    pivot_table.AddDataField(pivot_table.PivotFields('Total Order'),
                             'Count of HAWB', -4157)            # xlSum
    pivot_table.AddDataField(pivot_table.PivotFields('PBI Count'),
                             'Sum of PBI Count', -4157)

    # 계산 필드 C = 실패건수
    pivot_table.CalculatedFields().Add(
        Name="C",
        Formula="='Total Order' - 'PBI Count'"
    )
    calc_field = pivot_table.PivotFields("C")
    pivot_table.AddDataField(calc_field, "No of failed", -4157)

        # 계산 필드 D = On-time %
    # 0으로 나누는 경우 #DIV/0! 방지 → 조건부 수식 적용
    try:
        pivot_table.CalculatedFields().Item("D").Delete()
    except Exception:
        pass
    pivot_table.CalculatedFields().Add(
        Name="D",
        Formula="=IF('Total Order'=0,0,'PBI Count'/'Total Order')"
    )
    on_time = pivot_table.AddDataField(pivot_table.PivotFields("D"),
                                       "On Time Percentage", -4106)      # xlAverage
    on_time.NumberFormat = '0.00%'

    # ====== [NEW: Adjusted 로직 추가] ======
    try:
        # 1) Adjusted Count 합계
        pivot_table.AddDataField(pivot_table.PivotFields('Adjusted Count'),
                                 'Sum of Adjusted Count', -4157)   # xlSum

        # 2) Adjusted Failed = Total Order - Adjusted Count
        try:
            pivot_table.CalculatedFields().Item("AdjFailed").Delete()
        except Exception:
            pass
        pivot_table.CalculatedFields().Add(
            Name="AdjFailed",
            Formula="='Total Order' - 'Adjusted Count'"
        )
        adj_failed = pivot_table.PivotFields("AdjFailed")
        pivot_table.AddDataField(adj_failed, "Adjusted Failed", -4157)

        # 3) Adjusted On Time % = IF(Total Order=0,0,Adjusted Count/Total Order)
        try:
            pivot_table.CalculatedFields().Item("AdjOnTime").Delete()
        except Exception:
            pass
        pivot_table.CalculatedFields().Add(
            Name="AdjOnTime",
            Formula="=IF('Total Order'=0,0,'Adjusted Count'/'Total Order')"
        )
        adj_on_time = pivot_table.AddDataField(pivot_table.PivotFields("AdjOnTime"),
                                               "Adjusted On Time Percentage", -4106)  # xlAverage
        adj_on_time.NumberFormat = '0.00%'
    except Exception as e:
        print(f"[WARN] Adjusted fields creation failed: {e}")
    # =======================================



    wb.Sheets("PivotTable").Move(Before=wb.Sheets(1))

    # Move "All Data" sheet to second position
    wb.Sheets("Data_All").Move(Before=wb.Sheets(2))
    # ----- 저장 & 종료 -----
    wb.Save()
    wb.Close(SaveChanges=True)
    excel.Quit()
    pythoncom.CoUninitialize()
# ***************************************

# ----[추가: 헤더매핑 파싱 함수]----
def parse_header_mapping(xlsx_path, mapping_group):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 헤더군 리스트: E열(5번째)~오른쪽, 3번째 행(row=3)
    mapping_groups = []
    col = 5  # 1-indexed
    while ws.cell(row=3, column=col).value:
        val = ws.cell(row=3, column=col).value
        if not str(val).strip(): break
        mapping_groups.append(str(val).strip())
        col += 1

    # 찾고자 하는 매핑군의 컬럼 인덱스
    try:
        idx = mapping_groups.index(mapping_group)
    except ValueError:
        raise ValueError(f"Selected mapping group '{mapping_group}' not found in header mapping file.")

    # 표준헤더는 B열(2번째) 4번째 행~끝(예시, 25줄)
    std_headers = []
    std_row = 4
    while ws.cell(row=std_row, column=2).value:
        val = ws.cell(row=std_row, column=2).value
        if not str(val).strip(): break
        std_headers.append(str(val).strip())
        std_row += 1

    # 매핑 딕셔너리 만들기 (key: 실제파일의 컬럼, value: 표준컬럼)
    mapping_dict = {}
    for i, std_col in enumerate(std_headers):
        mapped_col = ws.cell(row=4+i, column=5+idx).value
        if mapped_col and str(mapped_col).strip():
            mapping_dict[str(mapped_col).strip()] = std_col

    # ====== [NEW: 정규화 매칭 지원 추가] ======
    def _normalize(s: str) -> str:
        return "".join(str(s).split()).lower() if s else ""

    # 원본 매핑 dict 유지, 대신 lookup 단계에서 정규화된 dict 병행 사용
    normalized_dict = {_normalize(k): v for k, v in mapping_dict.items()}

    # 반환 객체를 tuple로 구성 → (원본 dict, 정규화 dict)
    return {"original": mapping_dict, "normalized": normalized_dict}
# ----------------------------------

def allowed_columns(xlsx_path):
    df = pd.read_excel(xlsx_path, skiprows=2)
    allowed_names = df["Standardized Column"].iloc[0:].dropna().tolist()
    return allowed_names
# ----[추가: 이중헤더 보조 함수]----
def is_double_header(xlsx_path, sheet_name=None):
    """A2 값이 비어 있으면 두 줄 헤더로 판단"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    a2_val = ws['A2'].value
    return a2_val is None or str(a2_val).strip() == ""

def combine_header(xlsx_path, sheet_name=None):
    """두 줄 헤더를 한 줄로 병합한 DataFrame 반환"""

    # 1) 헤더 두 줄만 먼저 읽기
    headers_df_raw = pd.read_excel(
        xlsx_path, nrows=2, header=None, sheet_name=sheet_name
    )

    # ----[패치] pandas가 dict를 돌려줄 때 대비 -----------------
    if isinstance(headers_df_raw, dict):          # sheet_name=None ⇒ dict
        headers_df = next(iter(headers_df_raw.values()))
    else:
        headers_df = headers_df_raw
    # ----------------------------------------------------------

    skip = 2  # 데이터 시작 행

    # 2) Consignee/Consignor 중복 보정
    first_header  = headers_df.iloc[0].copy()
    second_header = headers_df.iloc[1].copy()
    duplicated_names = ["Consignee", "Consignor"]

    for name in duplicated_names:
        match_indices = second_header[second_header == name].index.tolist()
        if len(match_indices) >= 2:
            first_header.iloc[match_indices[0]]  = name
            second_header.iloc[match_indices[1]] = "Name"

    headers_df = pd.DataFrame([first_header, second_header])

    # 3) 빈 칸 채우고 두 줄 합치기
    headers_df.iloc[0] = headers_df.iloc[0].ffill()
    combined_headers = (
        headers_df.iloc[0].fillna("").astype(str).str.strip() + " " +
        headers_df.iloc[1].fillna("").astype(str).str.strip()
    ).str.strip()

    # 4) 본 데이터 읽기 & 헤더 적용
    df_raw = pd.read_excel(
        xlsx_path, skiprows=skip, header=None, sheet_name=sheet_name
    )
    if isinstance(df_raw, dict):
        df = next(iter(df_raw.values()))
    else:
        df = df_raw

    df.columns = combined_headers
    return df

# ----------------------------------

# ----[NEW: 중복 컬럼 Coalesce]---------------------------------
# def coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
#     """
#     동일한 컬럼명이 두 개 이상 있으면
#     • 왼쪽→오른쪽 순서로 값을 채워 넣고
#     • 최종적으로 컬럼을 하나만 남긴다.
#     """
#     dup_names = df.columns[df.columns.duplicated(keep=False)]
#     for name in dup_names.unique():
#         same_cols = df.loc[:, df.columns == name]
#         # NaN인 셀은 오른쪽 열 값으로 채움
#         df[name] = same_cols.bfill(axis=1).iloc[:, 0]
#     # 첫 번째 열만 남기고 나머지 삭제
#     return df.loc[:, ~df.columns.duplicated(keep='first')]
# ----------------------------------------------------------------

@app.route('/')
def upload_view():
    # 자동으로 폴더 이름을 기준으로 국가 리스트 생성
    steps_root = os.path.join(os.path.dirname(__file__), 'configured_steps')
    countries = sorted([
        name for name in os.listdir(steps_root)
        if os.path.isdir(os.path.join(steps_root, name))
    ])
    return render_template('index.html', countries=countries)

@app.route("/upload", methods=["POST"])
def handle_upload():
    # ── 0. 입력 수집 ──────────────────────────────────────
    raw_files  = request.files.getlist("raw_file[]")
    countries  = request.form.getlist("country[]")
    hawb_file  = request.files["hawb_file"]

    header_files          = request.files.getlist("header_file[]")
    header_mapping_groups = request.form.getlist("header_mapping_group[]")

    year  = int(request.form["target_year"])
    month = int(request.form["target_month"])

    # 길이 불일치 검증
    if len(raw_files) != len(countries):
        return "Raw 파일 수와 Country 선택 수가 일치하지 않습니다.", 400

    # ── 1. HAWB Validation 소스 저장 ─────────────────────
    tmp_hawb_path = os.path.join(
        UPLOAD_FOLDER, f"hawb_{uuid.uuid4().hex}.xlsx")
    hawb_file.save(tmp_hawb_path)

    # ── 2. 각 Raw 파일 처리 ───────────────────────────────
    results = []                      # (country, df) 튜플 목록
    for idx, (f, country) in enumerate(zip(raw_files, countries)):
        tmp_raw_path = os.path.join(
            UPLOAD_FOLDER, f"{country}_{uuid.uuid4().hex}.xlsx")
        f.save(tmp_raw_path)

        # ----[추가: 이중헤더 검사 & 병합]----
        if is_double_header(tmp_raw_path):
            df_merged = combine_header(tmp_raw_path)
            df_merged.to_excel(tmp_raw_path, index=False)

        # ── 2-A. 헤더 매핑(파일별) ────────────────────────
        header_mapping_dict = None
        if idx < len(header_files):
            header_file = header_files[idx]
            header_group = header_mapping_groups[idx] if idx < len(header_mapping_groups) else None
            if header_file and header_file.filename and header_group:
                header_map_path = os.path.join(
                    UPLOAD_FOLDER, f"header_map_{uuid.uuid4().hex}.xlsx")
                header_file.save(header_map_path)
                header_mapping_dict = parse_header_mapping(header_map_path, header_group)

        if header_mapping_dict:
            df_tmp = pd.read_excel(tmp_raw_path)
        
            # ★ 정규화 적용: df의 컬럼과 매핑파일 컬럼 비교 시 소문자+공백제거 기준
            def _normalize(s: str) -> str:
                return "".join(str(s).split()).lower() if s else ""
        
            normalized_map = header_mapping_dict["normalized"]
            original_map   = header_mapping_dict["original"]
        
            new_cols = {}
            for col in df_tmp.columns:
                norm_col = _normalize(col)
                if norm_col in normalized_map:
                    # 실제 rename 시에는 원본 dict 기준 값 사용 (예: HAWB)
                    new_cols[col] = normalized_map[norm_col]
        
            if new_cols:                       
                df_tmp.rename(columns=new_cols, inplace=True)
                df_tmp.to_excel(tmp_raw_path, index=False)


        # ----[추가: UPS 파일 Carrier 열 자동 추가]----
        # if "UPS" in f.filename.upper():
        #     df_carrier = pd.read_excel(tmp_raw_path)
        #     if "Carrier" not in df_carrier.columns:
        #         df_carrier["Carrier"] = "UPS"
        #         df_carrier.to_excel(tmp_raw_path, index=False)

        # ── 2-B. Validation 실행 ─────────────────────────
        df_ok = run_validation(
            raw_file_path  = tmp_raw_path,
            country        = country,
            hawb_file_path = tmp_hawb_path,
            year           = year,
            month          = month
        )
        df_ok["Country"] = country
        

        results.append((country, df_ok))

    # ── 3. 국가별 시트만 저장 ─────────────────────────────
    out_path = os.path.join(
        OUTPUT_FOLDER, f"validated_{year}_{month}_{uuid.uuid4().hex}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

# ***** [NEW] 모든 국가를 합친 시트 추가 *****
        if results:
            combined_df = pd.concat([d for _, d in results], ignore_index=True)
            # # ▼▼▼ (추가) HAWB 열이 없으면 UPS Trk를 복사
            # if "HAWB" not in combined_df.columns and "UPS Trk" in combined_df.columns:
            #     combined_df["HAWB"] = combined_df["UPS Trk"]
            # # 2) 열은 있는데 셀이 비어-있을 때 채우기  ← ★ 추가 3줄
            # if "HAWB" in combined_df.columns and "UPS Trk" in combined_df.columns:
            #     mask_blank = combined_df["HAWB"].isna() | (combined_df["HAWB"].astype(str).str.strip() == "")
            #     combined_df.loc[mask_blank, "HAWB"] = combined_df.loc[mask_blank, "UPS Trk"]
            
# ▼ HAWB → 없으면 UPS Trk 로 대체
            if "HAWB" in combined_df.columns:
                hawb_col = "HAWB"
            elif "UPS Trk" in combined_df.columns:
                hawb_col = "UPS Trk"
            else:
                hawb_col = None          # 둘 다 없으면 None

            if hawb_col:
                combined_df["HAWB_count_helper"] = combined_df[hawb_col].notna().astype(int)
            else:
                print("[WARN] HAWB / UPS Trk 컬럼이 없어 HAWB_count_helper 생성을 건너뜀")
                
            # ***** [NEW: Target Month] 사용자가 선택한 연·월로 엑셀 날짜 컬럼 추가 *****
            # pandas Timestamp를 쓰면 openpyxl 엔진이 엑셀 날짜로 기록해줍니다.
            try:
                combined_df["Target Month"] = pd.Timestamp(year=year, month=month, day=1)
            except Exception as _e:
                print(f"[WARN] Target Month 생성 실패: {_e}")
            # ***************************************************************

            
        # ***********************************************
            # NEW: Drop unmapped columns (for Data_All only)
            if header_mapping_dict:
                allowed_cols = allowed_columns(header_map_path)
                # Also include post-validation columns to retain in Data_All
                #additional = ["PBI", "FG", "WD", "Remarks", "Unique Order", "Total Order", "PBI Count", "Failed Order", "Adjusted Count", "Adjusted Reason", "On Time %", "Country","Target Month", "Target Year"]
                #allowed_cols += [col for col in additional if col in combined_df.columns]
                print(allowed_cols)
                combined_df = combined_df[[col for col in combined_df.columns if col in allowed_cols]]
                combined_df["Adjusted Count"] = combined_df["PBI Count"].copy()
                combined_df = combined_df.reindex(columns=allowed_cols)
            combined_df.to_excel(writer, index=False, sheet_name="Data_All")
        # ***********************************************
        used_names = set()                       # 시트명 중복 방지
        for country, df_res in results:
            # Excel 시트명 규칙: 31자 제한 + 특수문자 제거
            sheet = country[:31].translate(str.maketrans("/\\[]:*?", "_______"))
            # 중복 발생 시 _1, _2 … suffix 추가
            base = sheet or "Sheet"
            suffix = 1
            while sheet in used_names:
                sheet = f"{base[:30]}_{suffix}"
                suffix += 1
            used_names.add(sheet)

            df_res.to_excel(writer, index=False, sheet_name=sheet)

                # ***** [NEW] PivotTable 시트 삽입 *****
    try:
        add_pivot_to_workbook(out_path, data_sheet="Data_All")
    except Exception as e:
        print(f"[WARN] Pivot generation failed: {e}")
    # **************************************

    # ── 4. 결과 파일 다운로드 ──────────────────────────────
    return send_file(
        out_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="validated_result.xlsx"
    )

@app.route('/pipelines', methods=['GET', 'POST'])
def pipelines():
    if request.method == 'POST':
        country = request.form['country']
        pipeline_steps = request.form['pipeline_steps']
        import json
        try:
            steps = json.loads(pipeline_steps)
        except Exception:
            steps = []
        if country and steps:
            save_country_pipeline(country, steps)
        return redirect(url_for('pipelines'))

    all_steps = get_all_step_functions()
    pipelines = load_country_pipelines()
    return render_template('pipelines.html', all_steps=all_steps, pipelines=pipelines)

@app.route('/delete_pipeline', methods=['POST'])
def delete_pipeline():
    country = request.form['country']
    pipelines = load_country_pipelines()
    if country in pipelines:
        del pipelines[country]
        # 항상 절대경로로 저장!
    with open(path, 'r', encoding='utf-8') as f:
        step_obj = json.load(f)

    # 수정 전: return jsonify(step_obj)
    # 수정 후:
    return jsonify(step_obj["steps"])  # 배열만 반환



@app.route('/steps', methods=['GET'])
def get_steps():
    path = os.path.join(os.path.dirname(__file__), 'steps', 'unified_steps.json')

    if not os.path.exists(path):
        return jsonify({"error": f"File not found at path: {path}"}), 404

    try:
        with open(path, 'r', encoding='utf-8') as f:
            steps = json.load(f)
        print("[DEBUG] JSON loaded successfully.")
        return jsonify(steps)
    except Exception as e:
        print(f"[ERROR] Failed to load JSON: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/step-definition')
def get_step_definition():
    step_name = request.args.get('name')  # 예: step4_update_remarks

    if not step_name:
        return jsonify({'error': 'Missing step name'}), 400

    filename = f"{step_name}.json"
    filepath = os.path.join(os.path.dirname(__file__), 'step_definitions', filename)

    print(f"[DEBUG] Looking for file at: {filepath}")
    if not os.path.exists(filepath):
        print("[DEBUG] File does not exist.")
        return jsonify({'error': f'{filepath} not found'}), 404

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            step_def = json.load(f)
        print("[DEBUG] Step definition loaded successfully.")
        return jsonify(step_def)
    except Exception as e:
        print(f"[ERROR] Failed to load step definition: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/save-step', methods=['POST'])
def save_step():
    data = request.get_json() or {}

    country = data.get('country')
    step    = data.get('step')

    rules      = data.get('rules')       # 새 포맷
    conditions = data.get('conditions')  # 레거시 or 이미 감싼 포맷

    # ── 1. 필수값 검증 ─────────────────────────────────
    if not country or not step:
        return jsonify({"error": "Missing country or step"}), 400
    if rules is None and conditions is None:
        return jsonify({"error": "Missing rules or conditions"}), 400

    # ── 2. 저장 객체 생성 ──────────────────────────────
    if conditions is None:
        # rules만 왔을 때 → 규격에 맞게 감싸기
        save_obj = {
            "step": step,
            "conditions": { "rules": rules }
        }
    else:
        # 이미 conditions 키가 있으면 그대로 저장 (레거시 호환)
        save_obj = {
            "step": step,
            "conditions": conditions
        }

    # ── 3. 파일 저장 ──────────────────────────────────
    dir_path  = os.path.join(
        os.path.dirname(__file__), 'configured_steps', country)
    os.makedirs(dir_path, exist_ok=True)
    file_path = os.path.join(dir_path, f"{step}.json")

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(save_obj, f, indent=2, ensure_ascii=False)
        print(f"[DEBUG] Step saved to {file_path}")
        return jsonify({"message": "Step saved successfully"}), 200
    except Exception as e:
        print(f"[ERROR] Failed to save step: {e}")
        return jsonify({"error": str(e)}), 500

# # ========================
# # ***** [COMBINER ADDITIONS START] ***** We are no more using the Combiner feature-Jimin
# # ========================

# # ---- 추가 임포트 (중복 임포트 무해) ----
# import re
# import numpy as np
# from collections import defaultdict, Counter

# def _make_unique_sheet_name(existing: set, base: str) -> str:
#     """시트명 충돌 시 base(1), base(2) … 형태로 고유 이름 생성"""
#     if base not in existing:
#         existing.add(base)
#         return base
#     i = 1
#     while True:
#         cand = f"{base}({i})"
#         if cand not in existing:
#             existing.add(cand)
#             return cand
#         i += 1

# # ---------- 컬럼명 정규화 & 중복열 병합 유틸 ----------

# _ZWSP = "\u200b"  # zero width space 등 제거용

# def _pretty(s: str) -> str:
#     """표시용 예쁜 이름(최초 만난 원본 헤더를 최대한 보존)"""
#     return str(s).replace(_ZWSP, "").strip()

# def _canon(name: str) -> str:
#     """
#     '동일 의미' 판정을 위한 캐니컬 키:
#     - 끝의 '.1', '.2' 제거
#     - 공백/밑줄/구두점 제거
#     - 대문자화
#     - 제로폭문자 제거
#     """
#     s = "" if name is None else str(name)
#     s = s.replace(_ZWSP, "")
#     s = re.sub(r"\.\d+$", "", s)           # ".1" 접미사 정리
#     s = re.sub(r"\s+", " ", s).strip()     # 공백 normalize
#     s = re.sub(r"[^0-9A-Za-z]+", "", s)    # 비문자 제거
#     s = s.upper()
#     return s or "UNNAMED"

# def _coalesce_group(df: pd.DataFrame, cols: list[str]) -> pd.Series:
#     """
#     같은 의미(동일 캐니컬 키)의 여러 열을 하나로 병합:
#     - 공백 문자열/ "nan"/ "None" → 결측 취급
#     - 왼→오 순서로 첫 유효값 선택
#     """
#     merged = None
#     for c in cols:
#         s = df[c]
#         if s.dtype == object:
#             s = s.replace(r"^\s*$", pd.NA, regex=True)\
#                  .replace({"nan": pd.NA, "None": pd.NA})
#         if merged is None:
#             merged = s
#         else:
#             merged = merged.combine_first(s)
#     return merged

# def _clean_data_all_columns(df: pd.DataFrame):
#     """
#     1) 컬럼명 캐니컬화 → 그룹핑
#     2) 그룹별 coalesce
#     3) 완전 빈(전부 NA)이며 UNNAMED 류는 삭제
#     4) 리턴: (캐니컬-열 이름의 DF, 캐니컬→표시용 이름 맵)
#     """
#     # 모든 컬럼명을 문자열화
#     df = df.copy()
#     df.columns = [str(c) for c in df.columns]

#     groups = defaultdict(list)     # canon -> [original names...]
#     pretty = {}                    # canon -> first pretty name

#     for c in df.columns:
#         key = _canon(c)
#         groups[key].append(c)
#         if key not in pretty:
#             pretty[key] = _pretty(c)

#     data = {}
#     order = []                     # 출력 순서 보존
#     for key, cols in groups.items():
#         col = _coalesce_group(df, cols)
#         # 완전 빈 컬럼 & 이름이 의미없으면 버린다
#         if col.isna().all() and key.startswith("UNNAMED"):
#             continue
#         data[key] = col
#         order.append(key)

#     out = pd.DataFrame(data, columns=order)
#     return out, pretty

# def _union_ordered(keys_list: list[list[str]]) -> list[str]:
#     """여러 키 리스트를 '나온 순서대로' 유니온"""
#     seen, out = set(), []
#     for keys in keys_list:
#         for k in keys:
#             if k not in seen:
#                 seen.add(k); out.append(k)
#     return out

# def _choose_pretty_name(pretty_maps: list[dict], final_keys: list[str]) -> dict:
#     """
#     각 캐니컬 키에 대해 표시용 이름을 결정:
#     - 파일들의 pretty 후보 중 '가장 먼저 등장'한 걸 우선
#     - 동률이면 빈도 많은 이름 선택
#     """
#     result = {}
#     for k in final_keys:
#         candidates = [m[k] for m in pretty_maps if k in m]
#         if not candidates:
#             # 마지막 fallback: 캐니컬 키를 Title로 복원
#             result[k] = k.title()
#             continue
#         # 빈도 우선
#         cnt = Counter(candidates)
#         most = cnt.most_common()
#         top_count = most[0][1]
#         tops = [name for name, c in most if c == top_count]
#         # 등장 순서 우선
#         for cand in candidates:
#             if cand in tops:
#                 result[k] = cand
#                 break
#     return result

# def _safe_read_data_all(path: str) -> pd.DataFrame | None:
#     """
#     각 validated_result.xlsx의 Data_All을 '그대로' 읽어오되,
#     엑셀 개체/합치기/숨은문자 등 영향 최소화.
#     """
#     try:
#         df = pd.read_excel(path, sheet_name="Data_All", dtype=object, engine="openpyxl")
#         return df
#     except Exception as e:
#         print(f"[COMBINE] WARN: '{os.path.basename(path)}' Data_All 읽기 실패 → skip ({e})")
#         return None

# def _postprocess_types(df_final: pd.DataFrame, pretty_map: dict) -> pd.DataFrame:
#     """
#     대표적으로 'Target Month'는 날짜로 되살려 주고,
#     전부 NA인 열은 제거해 깔끔하게.
#     """
#     df = df_final.copy()

#     # 전체가 NA인 열 제거 (쓰레기 헤더 방지)
#     df = df.dropna(axis=1, how='all')

#     # Target Month 날짜 캐스팅 (있으면)
#     # 캐니컬 키를 모를 수 있으니 pretty 이름 기준도 함께 시도
#     possible_names = {"Target Month", "TARGET MONTH"}
#     for col in df.columns:
#         if str(col).strip() in possible_names:
#             try:
#                 df[col] = pd.to_datetime(df[col], errors="coerce")
#             except Exception:
#                 pass
#     return df

# # --------- v1(기존) 보존: 원래 함수명 변경해 둠 ----------
# def combine_validated_results_v1(file_paths: list[str]) -> str:
#     """
#     (보존용) 최초 버전: 단순 유니온-결합 + COM 시트복사 + Pivot 재생성
#     """
#     # 1) Data_All
#     frames = []
#     for p in file_paths:
#         df = _safe_read_data_all(p)
#         if df is not None:
#             frames.append(df)
#     if not frames:
#         raise ValueError("선택한 파일들에 유효한 Data_All 시트가 없습니다.")

#     # 유니온 순서
#     cols = []
#     for f in frames:
#         for c in f.columns:
#             if c not in cols:
#                 cols.append(c)
#     frames = [f.reindex(columns=cols) for f in frames]
#     combined_df = pd.concat(frames, ignore_index=True)

#     out_path = os.path.join(OUTPUT_FOLDER, f"combined_results_{uuid.uuid4().hex}.xlsx")
#     with pd.ExcelWriter(out_path, engine="openpyxl") as w:
#         combined_df.to_excel(w, index=False, sheet_name="Data_All")

#     # 2) Country 시트 복사
#     pythoncom.CoInitialize()
#     try:
#         makepy.GenerateFromTypeLibSpec("Microsoft Excel 16.0 Object Library")
#     except Exception:
#         pass
#     excel = Dispatch("Excel.Application"); excel.Visible = False
#     dest = excel.Workbooks.Open(os.path.abspath(out_path))
#     existing = set(s.Name for s in dest.Sheets)

#     for src_path in file_paths:
#         wb = excel.Workbooks.Open(os.path.abspath(src_path))
#         for sh in list(wb.Worksheets):
#             nm = sh.Name
#             if nm in ("Data_All", "PivotTable"):
#                 continue
#             sh.Copy(After=dest.Sheets(dest.Sheets.Count))
#             new_sh = dest.Sheets(dest.Sheets.Count)
#             unique = _make_unique_sheet_name(existing, nm)
#             try: new_sh.Name = unique
#             except Exception: pass
#         wb.Close(SaveChanges=False)

#     try: dest.Sheets("PivotTable").Delete()
#     except Exception: pass
#     dest.Close(SaveChanges=True); excel.Quit(); pythoncom.CoUninitialize()

#     try: add_pivot_to_workbook(out_path, data_sheet="Data_All")
#     except Exception as e: print(f"[COMBINE] Pivot generation failed: {e}")

#     return out_path

# # --------- v2(개선) : 정규화 + coalesce + 진단리포트 ----------
# def combine_validated_results_v2(file_paths: list[str]) -> str:
#     """
#     다수의 validated_result.xlsx → combined_results.xlsx (정확·깔끔 결합)
#       • 각 파일의 Data_All:
#           - 컬럼명 캐니컬화 → 중복열 coalesce → 의미없는 빈 열 제거
#       • 파일 간 유니온은 '등장 순서'를 지키며 수행
#       • 최종 Data_All은 전부 NA인 열 제거, Target Month 날짜 복원
#       • Country 시트는 원본 그대로 복사 (충돌명은 CR(1) 등으로 유니크화)
#       • 기존 PivotTable 삭제 후 통합 Data_All 기준 새 PivotTable 생성
#       • 마지막으로 Combine_Report 시트에 국가별 건수 대조 리포트 생성
#     """
#     cleaned_frames = []
#     pretty_maps = []
#     key_orders = []

#     # 1) 파일별 Data_All → 정규화/병합
#     for p in file_paths:
#         raw = _safe_read_data_all(p)
#         if raw is None:
#             continue
#         df_clean, pretty = _clean_data_all_columns(raw)
#         cleaned_frames.append(df_clean)
#         pretty_maps.append(pretty)
#         key_orders.append(list(df_clean.columns))

#     if not cleaned_frames:
#         raise ValueError("선택한 파일들에 유효한 Data_All 시트가 없습니다.")

#     # 2) 최종 컬럼 키 순서(등장 순서 유니온)
#     final_keys = _union_ordered(key_orders)

#     # 3) 표시용 이름 결정
#     final_pretty = _choose_pretty_name(pretty_maps, final_keys)

#     # 4) 유니온 정렬 후 행 결합
#     aligned = [f.reindex(columns=final_keys) for f in cleaned_frames]
#     combined = pd.concat(aligned, ignore_index=True)

#     # 5) 표시용 이름으로 리네임 + 타입 후처리
#     combined.rename(columns=final_pretty, inplace=True)
#     combined = _postprocess_types(combined, final_pretty)

#     # 6) 저장
#     out_path = os.path.join(OUTPUT_FOLDER, f"combined_results_{uuid.uuid4().hex}.xlsx")
#     with pd.ExcelWriter(out_path, engine="openpyxl") as w:
#         combined.to_excel(w, index=False, sheet_name="Data_All")

#     # 7) Country 시트 복사 (COM Excel)
#     pythoncom.CoInitialize()
#     try:
#         makepy.GenerateFromTypeLibSpec("Microsoft Excel 16.0 Object Library")
#     except Exception:
#         pass

#     excel = Dispatch("Excel.Application"); excel.Visible = False
#     dest = excel.Workbooks.Open(os.path.abspath(out_path))
#     existing_names = set(s.Name for s in dest.Sheets)

#     for src_path in file_paths:
#         wb = excel.Workbooks.Open(os.path.abspath(src_path))
#         for sh in list(wb.Worksheets):
#             nm = sh.Name
#             if nm in ("Data_All", "PivotTable"):
#                 continue
#             sh.Copy(After=dest.Sheets(dest.Sheets.Count))
#             new_sh = dest.Sheets(dest.Sheets.Count)
#             unique = _make_unique_sheet_name(existing_names, nm)
#             try: new_sh.Name = unique
#             except Exception: pass
#         wb.Close(SaveChanges=False)

#     # 8) 기존 Pivot 제거 후 새 PivotTable 생성
#     try: dest.Sheets("PivotTable").Delete()
#     except Exception: pass
#     dest.Close(SaveChanges=True); excel.Quit(); pythoncom.CoUninitialize()
#     try: add_pivot_to_workbook(out_path, data_sheet="Data_All")
#     except Exception as e: print(f"[COMBINE] Pivot generation failed: {e}")

#     # 9) 진단 리포트 시트 생성(국가별 건수 대조)
#     _write_combine_report(out_path)

#     return out_path

# def _write_combine_report(out_path: str) -> None:
#     """
#     통합 파일에 'Combine_Report' 시트를 추가:
#       - 각 국가 시트의 행 수
#       - Data_All의 Country별 행 수
#       - 차이(= 시트 행수 - Data_All 행수)
#     """
#     try:
#         # Data_All 카운트
#         df_all = pd.read_excel(out_path, sheet_name="Data_All", dtype=object, engine="openpyxl")
#         if "Country" in df_all.columns:
#             df_all_counts = (
#                 df_all
#                 .dropna(how="all")
#                 .assign(Country=df_all["Country"].astype(str).str.strip())
#                 .groupby("Country", dropna=False)
#                 .size()
#                 .reset_index(name="Rows in Data_All")
#                 .rename(columns={"Country": "Sheet"})
#             )
#         else:
#             df_all_counts = pd.DataFrame(columns=["Sheet", "Rows in Data_All"])

#         # 시트별 행수
#         wb = openpyxl.load_workbook(out_path, data_only=True)
#         country_sheets = [s for s in wb.sheetnames if s not in ("Data_All", "PivotTable")]
#         wb.close()

#         rows = []
#         for s in country_sheets:
#             try:
#                 d = pd.read_excel(out_path, sheet_name=s, dtype=object, engine="openpyxl")
#                 n = int(d.dropna(how="all").shape[0])
#             except Exception:
#                 n = 0
#             rows.append({"Sheet": s, "Rows in Sheet": n})
#         df_sheets = pd.DataFrame(rows)

#         report = pd.merge(df_sheets, df_all_counts, on="Sheet", how="left")
#         report["Rows in Data_All"] = report["Rows in Data_All"].fillna(0).astype(int)
#         report["Diff (Sheet - Data_All)"] = report["Rows in Sheet"] - report["Rows in Data_All"]

#         with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
#             report.to_excel(w, index=False, sheet_name="Combine_Report")
#     except Exception as e:
#         print(f"[COMBINE] report failed: {e}")

# @app.route('/combine', methods=['GET', 'POST'])
# def combine_view():
#     if request.method == 'GET':
#         return render_template('combine.html')

#     # POST
#     files = request.files.getlist('files[]')
#     if not files or len(files) < 2:
#         return "두 개 이상의 .xlsx 파일을 선택해 주세요.", 400

#     saved = []
#     for f in files:
#         if not f or not f.filename.lower().endswith('.xlsx'):
#             continue
#         path = os.path.join(UPLOAD_FOLDER, f"validated_{uuid.uuid4().hex}.xlsx")
#         f.save(path)
#         saved.append(path)

#     if len(saved) < 2:
#         return "두 개 이상의 유효한 .xlsx 파일이 필요합니다.", 400

#     try:
#         # *** 개선된 v2 로직 사용 (v1 보존됨) ***
#         out_path = combine_validated_results_v2(saved)
#     except Exception as e:
#         return f"병합 실패: {e}", 500

#     return send_file(
#         out_path,
#         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         as_attachment=True,
#         download_name="combined_results.xlsx"
#     )

# # ======================
# # ***** [COMBINER ADDITIONS END] *****
# # ======================


if __name__ == '__main__':
    app.run(debug=True)
